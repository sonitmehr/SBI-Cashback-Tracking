import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import csv
from helper import resolve_mode_from_csv
from raw_data.february_2026 import raw_input_json as input_json

YEAR = "2026"

# ---- Load JSON ----
data = json.loads(input_json)

entries = data["transactions"]
cashback = data["cashback"]

# ---- Month from first transaction ----
first_date = entries[0]["date"]          # "01 Feb 26"
_, month_label, _ = first_date.split()
MONTH_LABEL = month_label

# ---- Timestamped filename ----
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
FILE_NAME = f"statements/{MONTH_LABEL}_{YEAR}_{timestamp}.xlsx"
SHEET_NAME = "Transactions"

# ---- Headers ----
headers = [
    "Date",
    "Merchant Name",
    "Amount",
    "Type",
    "Mode",
    "Cashback Expected",
    "Net Payment",
    "Done By"
]

# ---- Load Merchant CSV ----
merchants_map = {}

print("\n📄 Reading merchants.csv\n")
with open("merchants.csv", newline="", encoding="utf-8") as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        merchants_map[row["Merchant Name"].strip()] = row["Mode"].strip().upper()

print("✅ Merchant mode mapping loaded\n")

# ---- Month Map ----
month_map = {
    "Jan": "01", "Feb": "02", "Mar": "03",
    "Apr": "04", "May": "05", "Jun": "06",
    "Jul": "07", "Aug": "08", "Sep": "09",
    "Oct": "10", "Nov": "11", "Dec": "12"
}

rows = []

# ---- Parse Entries ----
for entry in entries:

    day, month, year = entry["date"].split()

    merchant = entry["merchant"]
    amount = float(entry["amount"])
    txn_type = entry["type"]
    mode = entry["mode"]

    done_by = entry.get("done_by")

    # Default value if null / missing / empty
    if not done_by:
        done_by = "User"

    # Resolve ON/OFF/NO from CSV
    mode = resolve_mode_from_csv(
        entry=str(entry),
        merchant_name=merchant,
        current_mode=mode,
        merchant_mode_map=merchants_map
    )

    date = f"{YEAR}-{month_map[month]}-{day}"

    # Credit -> negative
    if txn_type == "C":
        amount = -abs(amount)
        txn_type = "Credit"
    else:
        amount = abs(amount)
        txn_type = "Debit"

    rows.append([
        date,
        merchant.strip(),
        amount,
        txn_type,
        mode,
        done_by
    ])

# ---- Create Workbook ----
wb = Workbook()
ws = wb.active
ws.title = SHEET_NAME
ws.append(headers)

# ---- Write Data ----
for i, r in enumerate(rows, start=2):
    ws.cell(i, 1, r[0])
    ws.cell(i, 2, r[1])
    ws.cell(i, 3, r[2])
    ws.cell(i, 4, r[3])
    ws.cell(i, 5, r[4])

    # Cashback rule
    ws.cell(
        i, 6,
        (
            f'=IF(ABS(C{i})<100,0,'
            f'FLOOR('
            f'ABS(C{i})*IF(E{i}="ON",0.05,IF(E{i}="OFF",0.01,0)),'
            f'1'
            f')*SIGN(C{i})'
            f')'
        )
    )

    ws.cell(i, 7, f"=C{i}-F{i}")
    ws.cell(i, 8, r[5])

# ---- Create Table ----
end_row = ws.max_row
table = Table(displayName="TransactionTable", ref=f"A1:H{end_row}")
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

# ---- Summary ----
summary_row = end_row + 2

ws.cell(summary_row, 2, "Total Transaction Amount")
ws.cell(summary_row, 3, f"=SUM(C2:C{end_row})")

ws.cell(summary_row + 1, 2, "Total Cashback Expected")
ws.cell(summary_row + 1, 3, f"=SUM(F2:F{end_row})")

ws.cell(summary_row + 2, 2, "Cashback Received")
ws.cell(summary_row + 2, 3, cashback)

ws.cell(summary_row + 3, 2, "Cashback Difference")
ws.cell(summary_row + 3, 3, f"=C{summary_row + 1}-C{summary_row + 2}")

# ---- Payment Pending Per User ----
users = sorted(set(r[5] for r in rows if r[5]))

start_col = 10
ws.cell(2, start_col, "Payment Pending From")

for idx, user in enumerate(users, start=1):
    row = 2 + idx
    ws.cell(row, start_col, user)
    ws.cell(
        row,
        start_col + 1,
        f'=SUMIF(H2:H{end_row},"{user}",G2:G{end_row})'
    )

# ---- Totals & Reconciliation ----
recon_start_row = 2 + len(users) + 2
recon_col = start_col

ws.cell(recon_start_row, recon_col, "Total Pending Payments")
ws.cell(
    recon_start_row,
    recon_col + 1,
    f"=SUM(K3:K{2 + len(users)})"
)

ws.cell(recon_start_row + 1, recon_col, "Cashback Received")
ws.cell(
    recon_start_row + 1,
    recon_col + 1,
    f"=C{summary_row + 2}"
)

ws.cell(recon_start_row + 2, recon_col, "Reconciliation Status")
ws.cell(
    recon_start_row + 2,
    recon_col + 1,
    (
        f'=IF('
        f'C{summary_row}=('
        f'K{recon_start_row}+K{recon_start_row + 1}'
        f'),'
        f'"Everything adds up",'
        f'"Something is not adding up"'
        f')'
    )
)

# ---- Save ----
wb.save(FILE_NAME)
print(f"✅ Excel generated successfully: {FILE_NAME}")