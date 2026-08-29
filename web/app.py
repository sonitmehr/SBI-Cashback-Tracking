from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
import json
import os
import csv
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename
from io import BytesIO
from helpers.utils import read_pdf_as_string, resolve_mode_from_csv, parse_statement, load_merchants_csv, get_mode_for_merchant, convert_display_mode_to_internal
try:
    from config import (
        get_merchants_from_db, save_merchant_for_approval, get_pending_approvals,
        approve_merchant, reject_merchant, initialize_merchants_from_csv, DEFAULT_ADMIN_PASSWORD
    )
    print("Using cloud MongoDB configuration")
except Exception as e:
    print(f"Cloud MongoDB config failed: {e}")
    print("Falling back to local MongoDB configuration")
    from config_local import (
        get_merchants_from_db, save_merchant_for_approval, get_pending_approvals,
        approve_merchant, reject_merchant, initialize_merchants_from_csv, DEFAULT_ADMIN_PASSWORD
    )

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

try:
    from dotenv import load_dotenv
    load_dotenv()
    parent_env = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')
    if os.path.exists(parent_env):
        load_dotenv(parent_env)
except ImportError:
    pass

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'sbi-cashback-tracking-secret-key-change-in-env')

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create upload directory if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Initialize merchants from CSV to MongoDB (one-time setup)
merchants_csv_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'merchants.csv')
initialize_merchants_from_csv(merchants_csv_path)

# Load merchants from MongoDB
merchants_map = get_merchants_from_db()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/get-merchants-map')
def get_merchants_map():
    """
    Return the merchants mapping for prefilling modes in the UI.
    """
    return jsonify({'merchants_map': merchants_map})


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    password = request.form.get('password', None)
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Parse the PDF with password if provided
            pdf_text = read_pdf_as_string(filepath, password)
            
            # Log raw PDF text
            print(f"\n{'='*80}")
            print(f"Raw PDF text:")
            print(f"{'='*80}")
            print(pdf_text)
            print(f"{'='*80}\n")
            
            parsed_data = parse_statement(pdf_text)
            
            # Log parsed data
            print(f"\n{'='*80}")
            print(f"Parsed PDF data:")
            print(f"{'='*80}")
            print(json.dumps(parsed_data, indent=2))
            print(f"{'='*80}\n")
            
            # Clean up uploaded file
            os.remove(filepath)
            
            return jsonify({
                'success': True,
                'data': parsed_data
            })
        except Exception as e:
            # Clean up uploaded file on error
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'Error parsing PDF: {str(e)}'}), 500
    
    return jsonify({'error': 'Invalid file type. Please upload a PDF.'}), 400


@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    try:
        data = request.json
        transactions = data.get('transactions', [])
        cashback = data.get('cashback', 0)
        
        if not transactions:
            return jsonify({'error': 'No transactions provided'}), 400
        
        # Get month from first transaction
        first_txn = transactions[0]
        txn_date = first_txn.get('Transaction date', '')
        if txn_date:
            parts = txn_date.split()
            if len(parts) >= 2:
                month_label = parts[1]
            else:
                month_label = 'Unknown'
        else:
            month_label = 'Unknown'
        
        YEAR = "2026"
        SHEET_NAME = "Transactions"
        
        headers = [
            "Date",
            "Merchant Name",
            "Amount",
            "Type",
            "Mode",
            "Cashback Expected",
            "Net Payment",
            "Done By"
        ]
        
        month_map = {
            "Jan": "01", "Feb": "02", "Mar": "03",
            "Apr": "04", "May": "05", "Jun": "06",
            "Jul": "07", "Aug": "08", "Sep": "09",
            "Oct": "10", "Nov": "11", "Dec": "12"
        }
        
        rows = []
        
        for txn in transactions:
            txn_date = txn.get('Transaction date', '')
            merchant = txn.get('Transaction Name', '')
            amount = float(txn.get('Amount', 0))
            txn_type = txn.get('Type', 'D')
            done_by = txn.get('done_by', 'User')
            display_mode = txn.get('mode', '')
            
            # Parse date
            parts = txn_date.split()
            if len(parts) >= 3:
                day, month, year = parts[0], parts[1], parts[2]
                date = f"{YEAR}-{month_map.get(month, '01')}-{day}"
            else:
                date = txn_date
            
            # Convert display mode (Online/Offline/Not Eligible) to internal mode (ON/OFF/NO)
            mode = convert_display_mode_to_internal(display_mode) if display_mode else None
            
            # If mode is still None, try to resolve from CSV as fallback
            if not mode:
                mode = resolve_mode_from_csv(merchant, None, merchants_map)
            
            # Credit -> negative
            if txn_type == 'C':
                amount = -abs(amount)
                txn_type = "Credit"
            else:
                amount = abs(amount)
                txn_type = "Debit"
            
            rows.append([
                date,
                merchant.strip(),
                amount,
                txn_type,
                mode,
                done_by
            ])
        
        # Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(headers)
        
        # Write Data
        for i, r in enumerate(rows, start=2):
            ws.cell(i, 1, r[0])
            ws.cell(i, 2, r[1])
            ws.cell(i, 3, r[2])
            ws.cell(i, 4, r[3])
            ws.cell(i, 5, r[4])
            
            # Cashback rule
            ws.cell(
                i, 6,
                (
                    f'=IF(ABS(C{i})<100,0,'
                    f'FLOOR('
                    f'ABS(C{i})*IF(E{i}="ON",0.05,IF(E{i}="OFF",0.01,0)),'
                    f'1'
                    f')*SIGN(C{i})'
                    f')'
                )
            )
            
            ws.cell(i, 7, f"=C{i}-F{i}")
            ws.cell(i, 8, r[5])
        
        # Create Table
        end_row = ws.max_row
        table = Table(displayName="TransactionTable", ref=f"A1:H{end_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Summary
        summary_row = end_row + 2
        
        ws.cell(summary_row, 2, "Total Transaction Amount")
        ws.cell(summary_row, 3, f"=SUM(C2:C{end_row})")
        
        ws.cell(summary_row + 1, 2, "Total Cashback Expected")
        ws.cell(summary_row + 1, 3, f"=SUM(F2:F{end_row})")
        
        ws.cell(summary_row + 2, 2, "Cashback Received")
        ws.cell(summary_row + 2, 3, cashback)
        
        ws.cell(summary_row + 3, 2, "Cashback Difference")
        ws.cell(summary_row + 3, 3, f"=C{summary_row + 1}-C{summary_row + 2}")
        
        # Payment Pending Per User
        users = sorted(set(r[5] for r in rows if r[5]))
        
        start_col = 10
        ws.cell(2, start_col, "Payment Pending From")
        
        for idx, user in enumerate(users, start=1):
            row = 2 + idx
            ws.cell(row, start_col, user)
            ws.cell(
                row,
                start_col + 1,
                f'=SUMIF(H2:H{end_row},"{user}",G2:G{end_row})'
            )
        
        # Totals & Reconciliation
        recon_start_row = 2 + len(users) + 2
        recon_col = start_col
        
        ws.cell(recon_start_row, recon_col, "Total Pending Payments")
        ws.cell(
            recon_start_row,
            recon_col + 1,
            f"=SUM(K3:K{2 + len(users)})"
        )
        
        ws.cell(recon_start_row + 1, recon_col, "Cashback Received")
        ws.cell(
            recon_start_row + 1,
            recon_col + 1,
            f"=C{summary_row + 2}"
        )
        
        ws.cell(recon_start_row + 2, recon_col, "Reconciliation Status")
        ws.cell(
            recon_start_row + 2,
            recon_col + 1,
            (
                f'=IF('
                f'C{summary_row}=('
                f'K{recon_start_row}+K{recon_start_row + 1}'
                f'),'
                f'"Everything adds up",'
                f'"Something is not adding up"'
                f')'
            )
        )
        
        # Save to BytesIO for in-memory download
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Send file directly to browser
        return send_file(
            output,
            as_attachment=True,
            download_name='statement.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error generating Excel: {str(e)}'}), 500


@app.route('/save-merchant-changes', methods=['POST'])
def save_merchant_changes():
    """
    Save merchant changes for admin approval.
    """
    try:
        data = request.get_json()
        changes = data.get('changes', [])
        user_ip = request.remote_addr
        
        saved_count = 0
        for change in changes:
            merchant_name = change.get('merchant_name')
            mode = change.get('mode')
            action_type = change.get('action_type', 'new')
            
            if merchant_name and mode:
                if save_merchant_for_approval(merchant_name, mode, action_type, user_ip):
                    saved_count += 1
        
        return jsonify({
            'success': True,
            'message': f'{saved_count} merchant changes saved for admin approval'
        })
    
    except Exception as e:
        return jsonify({'error': f'Error saving merchant changes: {str(e)}'}), 500


@app.route('/admin')
def admin_login():
    """
    Admin login page.
    """
    if 'admin_logged_in' in session:
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html')


@app.route('/admin/login', methods=['POST'])
def admin_authenticate():
    """
    Authenticate admin user.
    """
    password = request.form.get('password')
    if password == DEFAULT_ADMIN_PASSWORD:
        session['admin_logged_in'] = True
        return redirect(url_for('admin_dashboard'))
    else:
        return render_template('admin_login.html', error='Invalid password')


@app.route('/admin/dashboard')
def admin_dashboard():
    """
    Admin dashboard for approving merchant changes.
    """
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    pending_approvals = get_pending_approvals()
    return render_template('admin_dashboard.html', approvals=pending_approvals)


@app.route('/admin/approve/<approval_id>')
def admin_approve(approval_id):
    """
    Approve a merchant change.
    """
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    from bson import ObjectId
    if approve_merchant(ObjectId(approval_id)):
        # Refresh merchants map
        global merchants_map
        merchants_map = get_merchants_from_db()
        return jsonify({'success': True, 'message': 'Merchant approved successfully'})
    else:
        return jsonify({'error': 'Failed to approve merchant'}), 500


@app.route('/admin/reject/<approval_id>')
def admin_reject(approval_id):
    """
    Reject a merchant change.
    """
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    from bson import ObjectId
    if reject_merchant(ObjectId(approval_id)):
        return jsonify({'success': True, 'message': 'Merchant rejected successfully'})
    else:
        return jsonify({'error': 'Failed to reject merchant'}), 500


@app.route('/admin/bulk-action', methods=['POST'])
def admin_bulk_action():
    """
    Perform bulk approve or reject on multiple approvals.
    """
    if 'admin_logged_in' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json() or {}
    approval_ids = data.get('ids', [])
    action = data.get('action')  # 'approve' or 'reject'
    
    if not approval_ids or action not in ['approve', 'reject']:
        return jsonify({'error': 'Invalid request parameters'}), 400
    
    from bson import ObjectId
    success_count = 0
    fail_count = 0
    
    for aid in approval_ids:
        try:
            oid = ObjectId(aid)
            if action == 'approve':
                if approve_merchant(oid):
                    success_count += 1
                else:
                    fail_count += 1
            elif action == 'reject':
                if reject_merchant(oid):
                    success_count += 1
                else:
                    fail_count += 1
        except Exception as e:
            fail_count += 1
            
    if action == 'approve' and success_count > 0:
        global merchants_map
        merchants_map = get_merchants_from_db()
        
    action_verb = "approved" if action == "approve" else "rejected"
    return jsonify({
        'success': True,
        'message': f'Successfully {action_verb} {success_count} item(s)' + (f' ({fail_count} failed)' if fail_count > 0 else '')
    })


@app.route('/admin/logout')
def admin_logout():
    """
    Logout admin user.
    """
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login'))


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
