#!/usr/bin/env python3
"""
Sync Excel Merchants Script
Parses 'SBI Cashback Card & PhonePe Purple_Black.xlsx' and submits diffs
(new merchants and mode updates) to MongoDB for Admin Approval.
Designed to run locally and via GitHub Actions on a weekly schedule.
"""

import os
import sys
from datetime import datetime
from collections import Counter
import openpyxl

# Force UTF-8 output
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# Load environment variables if running locally
try:
    from dotenv import load_dotenv
    load_dotenv()
    parent_env = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(parent_env):
        load_dotenv(parent_env)
except ImportError:
    pass

from pymongo import MongoClient

# Configuration
EXCEL_FILE_NAME = (os.environ.get("EXCEL_FILE_PATH") or "SBI Cashback Card & PhonePe Purple_Black.xlsx").strip()
SHEET_NAME = (os.environ.get("EXCEL_SHEET_NAME") or "SBI Cashback").strip()
MONGODB_URI = os.environ.get("MONGODB_URI", "").strip()
DATABASE_NAME = (os.environ.get("DATABASE_NAME") or "sbi_cashback_tracking").strip()
MERCHANTS_COLLECTION = (os.environ.get("MERCHANTS_COLLECTION") or "merchants").strip()
ADMIN_COLLECTION = (os.environ.get("ADMIN_COLLECTION") or "admin_approvals").strip()


def map_cashback_to_mode(cb_val):
    """
    Map Excel 'Cashback received' value to internal mode (ON/OFF/NO).
    0.05 / 5% -> ON (Online 5%)
    0.01 / 1% -> OFF (Offline 1%)
    0.0  / 0% -> NO (Not Eligible 0%)
    """
    if cb_val is None:
        return None
    
    if isinstance(cb_val, (int, float)):
        if abs(cb_val - 0.05) < 0.005 or abs(cb_val - 5.0) < 0.1:
            return "ON"
        elif abs(cb_val - 0.01) < 0.005 or abs(cb_val - 1.0) < 0.1:
            return "OFF"
        elif abs(cb_val - 0.0) < 0.005:
            return "NO"
        return None

    cb_str = str(cb_val).strip().upper()
    if cb_str in ["5%", "0.05", "5", "5.0", "ON", "ONLINE"]:
        return "ON"
    elif cb_str in ["1%", "0.01", "1", "1.0", "OFF", "OFFLINE"]:
        return "OFF"
    elif cb_str in ["0%", "0.0", "0", "NO", "NOT ELIGIBLE", "NONE"]:
        return "NO"
    return None


def parse_merchants_from_excel(excel_path, sheet_name="SBI Cashback"):
    """
    Extracts unique merchants and their mapped modes from the Excel file.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")

    print(f"📖 Opening Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    ws = wb[sheet_name]
    print(f"📄 Reading sheet: '{sheet_name}' (Total Rows: {ws.max_row})")

    # Locate headers row
    header_row_idx = None
    merchant_col_idx = None
    cashback_col_idx = None

    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, min(20, ws.max_column + 1))]
        if "Merchant name on statement" in row_vals:
            header_row_idx = r
            merchant_col_idx = row_vals.index("Merchant name on statement") + 1
            if "Cashback received" in row_vals:
                cashback_col_idx = row_vals.index("Cashback received") + 1
            break

    if not header_row_idx or not merchant_col_idx or not cashback_col_idx:
        raise ValueError("Could not find required columns ('Merchant name on statement', 'Cashback received')")

    print(f"✅ Found header row at line {header_row_idx}: Merchant Col={merchant_col_idx}, Cashback Col={cashback_col_idx}")

    excel_merchants = {}
    skipped_unmapped = 0

    for r in range(header_row_idx + 1, ws.max_row + 1):
        raw_name = ws.cell(r, merchant_col_idx).value
        raw_cb = ws.cell(r, cashback_col_idx).value

        if not raw_name or not str(raw_name).strip():
            continue

        merchant_name = str(raw_name).strip()
        mode = map_cashback_to_mode(raw_cb)

        if not mode:
            skipped_unmapped += 1
            continue

        # If duplicate in Excel, later row overwrites or keeps last confirmed
        excel_merchants[merchant_name] = mode

    print(f"📊 Extracted {len(excel_merchants)} valid unique merchants from Excel (Skipped {skipped_unmapped} without valid mode)")
    return excel_merchants


def sync_with_mongodb(excel_merchants):
    """
    Computes diff against existing MongoDB collections and submits new / updated
    merchants to `admin_approvals` for admin review.
    """
    if not MONGODB_URI:
        raise ValueError("MONGODB_URI environment variable is missing!")

    print("🔌 Connecting to MongoDB Atlas...")
    client = MongoClient(
        MONGODB_URI,
        serverSelectionTimeoutMS=10000,
        connectTimeoutMS=10000,
        socketTimeoutMS=10000
    )
    client.admin.command('ping')
    print("✅ Connected to MongoDB Atlas successfully!")

    db = client[DATABASE_NAME]
    merchants_col = db[MERCHANTS_COLLECTION]
    admin_col = db[ADMIN_COLLECTION]

    # Fetch currently approved merchants in DB
    existing_merchants = {}
    for doc in merchants_col.find():
        name = doc.get("name", "").strip()
        mode = doc.get("mode", "").strip().upper()
        if name:
            existing_merchants[name.lower()] = {
                "original_name": name,
                "mode": mode
            }

    # Fetch existing pending approvals in DB to avoid duplicate submissions
    pending_approvals = {}
    for doc in admin_col.find({"status": "pending"}):
        p_name = doc.get("merchant_name", "").strip()
        p_mode = doc.get("mode", "").strip().upper()
        if p_name and p_mode:
            pending_approvals.setdefault(p_name.lower(), set()).add(p_mode)

    print(f"📦 Existing in DB: {len(existing_merchants)} approved merchants | {len(pending_approvals)} pending merchants with approvals")

    new_submissions = []
    update_submissions = []
    skipped_identical = 0
    skipped_already_pending = 0

    from datetime import timezone
    now = datetime.now(timezone.utc)

    for merchant_name, mode in excel_merchants.items():
        key = merchant_name.lower()

        # Check if already pending approval with exact same mode
        if key in pending_approvals and mode in pending_approvals[key]:
            skipped_already_pending += 1
            continue

        # Case 1: Merchant not in DB yet -> New Merchant
        if key not in existing_merchants:
            new_submissions.append({
                "merchant_name": merchant_name,
                "mode": mode,
                "action_type": "new",
                "status": "pending",
                "submitted_by": "weekly-excel-sync",
                "submitted_at": now,
                "approved_by": None,
                "approved_at": None
            })
        # Case 2: Merchant exists in DB but Mode is different -> Update Mode
        elif existing_merchants[key]["mode"] != mode:
            update_submissions.append({
                "merchant_name": existing_merchants[key]["original_name"],
                "mode": mode,
                "action_type": "update",
                "status": "pending",
                "submitted_by": "weekly-excel-sync",
                "submitted_at": now,
                "approved_by": None,
                "approved_at": None
            })
        else:
            # Case 3: Exists in DB with exact same mode -> No change
            skipped_identical += 1

    total_diffs = len(new_submissions) + len(update_submissions)

    print("\n" + "=" * 55)
    print("📈 SYNC SUMMARY:")
    print(f"  • Total Valid Merchants in Excel: {len(excel_merchants)}")
    print(f"  • Identical to DB (Up to Date):  {skipped_identical}")
    print(f"  • Already Pending in Approvals:  {skipped_already_pending}")
    print(f"  • New Merchants to Submit:       {len(new_submissions)}")
    print(f"  • Mode Updates to Submit:        {len(update_submissions)}")
    print("=" * 55)

    if total_diffs > 0:
        docs_to_insert = new_submissions + update_submissions
        print(f"\n🚀 Submitting {len(docs_to_insert)} change(s) to 'admin_approvals' collection...")
        admin_col.insert_many(docs_to_insert)
        print("✅ All changes successfully submitted for admin approval!")
        print("💡 Admin can review and approve them at /admin")
    else:
        print("\n✨ Database is completely in sync with Excel. No new changes to submit.")

    return total_diffs


def main():
    excel_path = os.path.join(os.path.dirname(__file__), EXCEL_FILE_NAME)
    if not os.path.exists(excel_path):
        excel_path = EXCEL_FILE_NAME

    try:
        excel_merchants = parse_merchants_from_excel(excel_path, SHEET_NAME)
        diff_count = sync_with_mongodb(excel_merchants)
        print("\n🎉 Sync process completed successfully!")
    except Exception as e:
        print(f"\n❌ Error during sync: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
